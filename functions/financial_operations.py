from tkinter import filedialog, Tk

import openpyxl

from data import mongodb_data
from functions.data_convert_operations import *


def create_excel_in_certain_period(start_date, end_date):
    # Fetch data from the database
    collection = mongodb_data.mongodb_instance.db["sales-0.1"]
    results = fetch_data_from_db(collection, start_date, end_date)

    # Sort the results by timestamp, and then by product name
    results.sort(key=lambda x: (x["timestamp"], x["product"]["name"]))

    # Create a new excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MySheet"

    # Create titles for columns
    titles = ["ID","Timestamp", "Product ID", "Name", "Price", "Profit Percentage", "Description", "Payment Method",
              "Vending Machine ID"]
    for col_num, title in enumerate(titles, 1):
        ws.cell(row=1, column=col_num, value=title)

    # Write data to the sheet
    for row_num, record in enumerate(results, 2):
        ws.cell(row=row_num, column=1, value=str(record['_id']))
        ws.cell(row=row_num, column=2, value=datetime.fromtimestamp(record["timestamp"]).strftime("%Y/%m/%d"))
        ws.cell(row=row_num, column=3, value=record["product"]["productID"])
        ws.cell(row=row_num, column=4, value=record["product"]["name"])
        ws.cell(row=row_num, column=5, value=record["product"]["price"])
        ws.cell(row=row_num, column=6, value=record["product"]["profit percentage"])
        ws.cell(row=row_num, column=7, value=record["product"]["description"])
        ws.cell(row=row_num, column=8, value=record["paymentMethod"])
        ws.cell(row=row_num, column=9, value=record["vendingMachineID"])

    # Use tkinter to choose save location
    root = Tk()
    root.withdraw()
    filename = filedialog.asksaveasfilename(
        initialfile=f"gegeven_van_{start_date.replace('/', '_')}_tot_{end_date.replace('/', '_')}.xlsx",
        defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])

    if filename:
        wb.save(filename)


def read_excel_and_convert_to_dict():
    # Create a tkinter root window and hide it
    root = Tk()
    root.withdraw()

    # Ask the user to select an Excel file
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])

    if file_path:
        # Load the Excel file
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active

        # Initialize a list to store the converted data
        data = []

        # Iterate through rows in the Excel sheet
        for row in ws.iter_rows(min_row=2, values_only=True):
            # Check if the row contains enough values
            if len(row) == 9:
                record = {
                    "_id": row[0],
                    "timestamp": row[1],
                    "product": {
                        "productID": row[2],
                        "name": row[3],
                        "price": row[4],
                        "profit percentage": row[5],
                        "description": row[6],
                    },
                    "paymentMethod": row[7],
                    "vendingMachineID": row[8],
                }

                # Append the record to the data list
                data.append(record)
            else:
                print(f"Skipping row with insufficient values: {row}")

        # Close the Excel file
        wb.close()

        return data
    else:
        return []